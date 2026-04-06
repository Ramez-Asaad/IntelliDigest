"""
document_loader.py
------------------
Multi-format document parser that extracts text from PDF, DOCX, Excel, and
plain text files. Includes case-fold normalization and semantic chunking.

Derived from Lab 2 — enhanced for production use.
"""

import os
from PyPDF2 import PdfReader
from docx import Document as DocxDocument
from openpyxl import load_workbook


# ── Extraction Functions ──────────────────────────────────────────────────────

def extract_text_from_pdf(file_path: str) -> str:
    """Extract all text from a PDF file."""
    reader = PdfReader(file_path)
    pages = [page.extract_text() or "" for page in reader.pages]
    return "\n".join(pages)


def extract_text_from_docx(file_path: str) -> str:
    """Extract all paragraph text from a DOCX file."""
    doc = DocxDocument(file_path)
    return "\n".join(para.text for para in doc.paragraphs if para.text.strip())


def extract_text_from_excel(file_path: str) -> str:
    """Extract cell values from the active sheet of an Excel file."""
    wb = load_workbook(file_path, data_only=True)
    ws = wb.active
    rows = []
    for row in ws.iter_rows(values_only=True):
        row_text = " | ".join(str(cell) if cell is not None else "" for cell in row)
        rows.append(row_text)
    return "\n".join(rows)


def extract_text_from_txt(file_path: str) -> str:
    """Read a plain text file."""
    with open(file_path, "r", encoding="utf-8", errors="ignore") as f:
        return f.read()


# ── File-Type Router ──────────────────────────────────────────────────────────

LOADERS = {
    ".pdf": extract_text_from_pdf,
    ".docx": extract_text_from_docx,
    ".xlsx": extract_text_from_excel,
    ".txt": extract_text_from_txt,
    ".md": extract_text_from_txt,
}

SUPPORTED_EXTENSIONS = list(LOADERS.keys())


def load_document(file_path: str) -> str:
    """
    Detect file type and extract text.

    Raises:
        ValueError: If the file extension is not supported.
    """
    ext = os.path.splitext(file_path)[1].lower()
    loader = LOADERS.get(ext)
    if loader is None:
        raise ValueError(
            f"Unsupported file type '{ext}'. "
            f"Supported: {', '.join(SUPPORTED_EXTENSIONS)}"
        )
    return loader(file_path)


def load_uploaded_file(uploaded_file) -> str:
    """
    Extract text from a Streamlit UploadedFile object.
    Writes to a temp file, then delegates to load_document.
    """
    import tempfile

    ext = os.path.splitext(uploaded_file.name)[1].lower()
    if ext not in LOADERS:
        raise ValueError(
            f"Unsupported file type '{ext}'. "
            f"Supported: {', '.join(SUPPORTED_EXTENSIONS)}"
        )

    with tempfile.NamedTemporaryFile(delete=False, suffix=ext) as tmp:
        tmp.write(uploaded_file.getvalue())
        tmp_path = tmp.name

    try:
        return load_document(tmp_path)
    finally:
        os.unlink(tmp_path)


# ── Text Processing ──────────────────────────────────────────────────────────

def normalize_text(text: str) -> str:
    """Case-fold normalize text for consistent processing."""
    return text.casefold()


def semantic_chunk(text: str, max_chunk_size: int = 500) -> list[str]:
    """
    Split text into semantically meaningful chunks by sentence boundaries,
    keeping each chunk under max_chunk_size characters.
    """
    sentences: list[str] = []
    current = ""

    for char in text:
        current += char
        if char in ".!?" and len(current.strip()) > 0:
            sentences.append(current.strip())
            current = ""

    if current.strip():
        sentences.append(current.strip())

    chunks: list[str] = []
    current_chunk = ""

    for sentence in sentences:
        if current_chunk and len(current_chunk) + len(sentence) > max_chunk_size:
            chunks.append(current_chunk.strip())
            current_chunk = sentence
        else:
            current_chunk += (" " + sentence) if current_chunk else sentence

    if current_chunk:
        chunks.append(current_chunk.strip())

    return chunks
